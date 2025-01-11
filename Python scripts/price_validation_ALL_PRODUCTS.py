import json
import pandas as pd
import requests
import openpyxl
from collections import defaultdict
from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def get_json_from_api(url):
    try:
        #response = requests.get(url)
        response = requests.get(url, verify=False)

        if response.status_code == 200:
            data = json.loads(response.content.decode('utf-8'))
        else:
            print(f"Unabale to get data from API, status code is {response.status_code}")
            data = None
    except requests.exceptions.RequestException as e:
        print(f"Error while sending API request {e}")
        data = None
        
    return data


def get_all_locations(api):
    all_locations = []
    for info in api['content']:
        locationId = info.get('id')
        location_status = info.get('status')
        if location_status == 'OPEN':
            all_locations.append(locationId)
    return all_locations

api_url = f"https://location-service-v2.api-idp.sonicdrivein.com/brand/SDI/location?size=10000"
api = get_json_from_api(api_url)

all_locations = get_all_locations(api)


def price_in_all_items(api, location):
    items_dict = {}
    size_id_to_name = {}

    for products_values in api['products'].values():
        for item_keys, item_values in products_values['items'].items():
            items_dict[item_values['id']] = item_values['name']

    for size_values in api['sizeGroups'].values():
        size_id_to_name[size_values['id']] = size_values['name']  

    if api is None:
        print("API data is None!")
        return None
    
    data = {'location': location}
    all_items_at_this_location = []
    for values in api['products'].values():
        product_name = values.get('name')
        product_type = values.get('type')
        if product_type == 'PRODUCT' and product_name != 'CONDIMENTS':
            for item_keys, item_values in values['items'].items():   
                item_size = item_values.get('sizeGroupId')
                all_items_at_this_location.append(item_keys)
                for i in all_items_at_this_location:
                    if item_keys == i:
                        item_price = item_values.get('price') 
                        size_name = size_id_to_name[item_size]
                        if size_name == 'NONE':
                            item_name = items_dict[i]
                        else:
                            item_name = f"{items_dict[i]} {size_name}"
                        if item_price:
                            #data[f"{items_dict[i]} {size_id_to_name[item_size]}"]  = item_price.get('currentPrice')
                            data[item_name] = item_price.get('currentPrice')
                        elif item_price == {}:
                            #data[f"{items_dict[i]} {size_id_to_name[item_size]}"]  = 'empty {}'
                            data[item_name] = 'Empty {}'
                        elif not item_price and item_price != {}:
                            #data[f"{items_dict[i]} {size_id_to_name[item_size]}"]  = None
                            data[item_name] = 'No Price'

    return data   

data_list = []
for location in all_locations:
    result = {}
    api_url1 = f"https://api-idp.sonicdrivein.com/snc/menu-api/menu/v1/brand/SDI/location/{location}/channel/WEBOA/type/ALLDAY"
    #api_url = f"https://api-idp.sonicdrivein.com/snc/web-exp-api/v1/menu/type/ALLDAY/id/{location}?sellingChannel=WEBOA"
    #api_url = f"https://menu-api-v0.snc-api.uat.irb.digital/menu/v1/brand/SDI/location/{location}/channel/WEBOA/type/ALLDAY"

    api_data = get_json_from_api(api_url1)
    if api_data is not None:
        result = price_in_all_items(api_data, location)
    else:
        result = {'location': location, 'price': 'Failed to get data for location'}
    data_list.append(result)

df = pd.DataFrame(data_list)
df.fillna('Not in API', inplace=True)
location_col = df.pop('location')
df = df.sort_index(axis=1)
df.insert(0, 'location', location_col)
df.to_excel('All_items_All_locations_01_11.xlsx', index=False)
#df = pd.DataFrame(data_list, columns=['location', 'item', 'price'])
#pivot_df = df.pivot(index='item', columns='location', values='price')
#pivot_df.to_excel('All_items_All_locations2.xlsx')
# Открываем файл с использованием openpyxl
wb = load_workbook('All_items_All_locations_01_11.xlsx')
ws = wb.active

# Change column width
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# Save changes
wb.save('All_items_All_locations_01_11.xlsx')

#print(price_at_all_items4(api_data, location))


from openpyxl import load_workbook

# open the workbook 
wb = load_workbook('All_items_All_locations_01_11.xlsx') 

# Select a particular sheet 
sheet = wb.active 

# Set the width of all columns in the sheet
for column in sheet.columns:
    max_length = 0
    column = column[0]
    sheet.column_dimensions[column.column_letter].width = 20

# Save the workbook
wb.save('All_items_All_locations_01_11.xlsx')